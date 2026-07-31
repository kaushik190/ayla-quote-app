import os
import re
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from kivy.lang import Builder
from kivy.utils import platform
from kivy.uix.image import Image
from kivy.uix.popup import Popup
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.filechooser import FileChooserListView
from kivymd.app import MDApp
from kivymd.uix.button import MDRaisedButton, MDFlatButton
from kivymd.uix.card import MDCard
from kivymd.uix.dialog import MDDialog
from kivymd.uix.gridlayout import MDGridLayout
from kivymd.uix.label import MDLabel
from kivymd.uix.textfield import MDTextField
from kivymd.uix.datatables import MDDataTable
from kivy.metrics import dp

KV = '''
BoxLayout:
    orientation: 'vertical'

    MDTopAppBar:
        title: "Product Quote Sheet"
        elevation: 1
        right_action_items: [["content-save", lambda x: app.prompt_excel_filename()], ["stop-circle-outline", lambda x: app.finish_session_prompt()]]

    MDScrollView:
        id: main_scroll
        MDBoxLayout:
            id: main_layout
            orientation: 'vertical'
            padding: "8dp"
            spacing: "8dp"
            size_hint_y: None
            height: self.minimum_height

            # Compact Form Card
            MDCard:
                orientation: 'vertical'
                padding: "10dp"
                spacing: "6dp"
                size_hint_y: None
                height: self.minimum_height
                elevation: 1

                # Top Row: Compact Photo + Description
                MDBoxLayout:
                    orientation: 'horizontal'
                    size_hint_y: None
                    height: "70dp"
                    spacing: "8dp"

                    Image:
                        id: image_preview
                        source: ''
                        size_hint: (None, None)
                        size: ("65dp", "65dp")
                        allow_stretch: True
                        keep_ratio: True

                    MDBoxLayout:
                        orientation: 'vertical'
                        spacing: "2dp"
                        size_hint_y: None
                        height: "65dp"

                        MDTextField:
                            id: field_desc
                            hint_text: "Item Description *"
                            size_hint_y: None
                            height: "36dp"
                            on_focus: if self.focus: app.scroll_to_widget(self)

                        MDBoxLayout:
                            orientation: 'horizontal'
                            spacing: "4dp"

                            MDRaisedButton:
                                text: "Camera"
                                icon: "camera"
                                font_size: "11sp"
                                size_hint_y: None
                                height: "30dp"
                                on_release: app.launch_native_camera(mode="item")

                            MDRaisedButton:
                                text: "Gallery"
                                icon: "folder-image"
                                font_size: "11sp"
                                size_hint_y: None
                                height: "30dp"
                                on_release: app.open_file_browser_popup()

                # 2-Column Compact Numerical Inputs
                MDGridLayout:
                    cols: 2
                    spacing: "6dp"
                    size_hint_y: None
                    height: self.minimum_height

                    MDTextField:
                        id: field_price
                        hint_text: "Price ($/¥)"
                        input_filter: 'float'
                        on_focus: if self.focus: app.scroll_to_widget(self)

                    MDTextField:
                        id: field_moq
                        hint_text: "MOQ"
                        input_filter: 'int'
                        on_focus: if self.focus: app.scroll_to_widget(self)

                    MDTextField:
                        id: field_qty_ctn
                        hint_text: "Qty / CTN"
                        input_filter: 'int'
                        on_focus: if self.focus: app.scroll_to_widget(self)

                    MDTextField:
                        id: field_cbm_ctn
                        hint_text: "CBM / CTN"
                        input_filter: 'float'
                        on_focus: if self.focus: app.scroll_to_widget(self)

                # Supplier Section Header with Scan Card Button
                MDBoxLayout:
                    orientation: 'horizontal'
                    size_hint_y: None
                    height: "32dp"
                    spacing: "8dp"

                    MDLabel:
                        text: "Supplier Info"
                        font_style: "Subtitle2"
                        theme_text_color: "Secondary"

                    MDRaisedButton:
                        text: "Scan Name Card"
                        icon: "card-account-details-outline"
                        font_size: "11sp"
                        size_hint_y: None
                        height: "28dp"
                        md_bg_color: 0.2, 0.6, 0.5, 1
                        on_release: app.launch_native_camera(mode="card")

                # Supplier Input Grid
                MDGridLayout:
                    cols: 2
                    spacing: "6dp"
                    size_hint_y: None
                    height: self.minimum_height

                    MDTextField:
                        id: field_shop
                        hint_text: "Shop / Booth No."
                        on_focus: if self.focus: app.scroll_to_widget(self)

                    MDTextField:
                        id: field_phone
                        hint_text: "Phone Number"
                        on_focus: if self.focus: app.scroll_to_widget(self)

                    MDTextField:
                        id: field_wechat
                        hint_text: "WeChat ID"
                        on_focus: if self.focus: app.scroll_to_widget(self)

                    MDTextField:
                        id: field_qq
                        hint_text: "QQ Number"
                        on_focus: if self.focus: app.scroll_to_widget(self)

                # Form Action Buttons
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: "8dp"
                    size_hint_y: None
                    height: "40dp"

                    MDRaisedButton:
                        text: "Add Item To Sheet"
                        icon: "plus-box"
                        on_release: app.save_quote()
                        size_hint_x: 0.65

                    MDRaisedButton:
                        text: "New Session"
                        icon: "stop"
                        md_bg_color: 0.8, 0.2, 0.2, 1
                        on_release: app.finish_session_prompt()
                        size_hint_x: 0.35

            # Live Preview Section Title
            MDLabel:
                text: "Live Excel Sheet Preview"
                font_style: "Subtitle2"
                size_hint_y: None
                height: "20dp"

            # Excel Live Data Table Container
            MDBoxLayout:
                id: table_container
                orientation: 'vertical'
                size_hint_y: None
                height: "250dp"
'''

class QuoteApp(MDApp):
    CAMERA_ITEM_REQUEST_CODE = 1024
    CAMERA_CARD_REQUEST_CODE = 1025

    def build(self):
        self.theme_cls.primary_palette = "Teal"
        self.current_photo_path = ""
        self.card_photo_path = ""
        self.quotes = []
        self.file_popup = None
        self.naming_dialog = None
        
        # Folder structure: Documents/QuoteSheets/ and Documents/QuoteSheets/Photos/
        base_doc_dir = "/sdcard/Documents" if os.path.exists("/sdcard/Documents") else os.path.expanduser("~/Documents")
        self.excel_dir = os.path.join(base_doc_dir, "QuoteSheets")
        self.photos_dir = os.path.join(self.excel_dir, "Photos")
        
        os.makedirs(self.excel_dir, exist_ok=True)
        os.makedirs(self.photos_dir, exist_ok=True)

        self.current_session_filename = ""
        
        root = Builder.load_string(KV)
        return root

    def on_start(self):
        self.init_data_table()
        if platform == 'android':
            try:
                from android import activity
                activity.bind(on_activity_result=self.on_activity_result)
            except Exception as e:
                print(f"Activity result bind failed: {e}")

    def scroll_to_widget(self, widget):
        """Scrolls the focused text field above soft keyboard."""
        self.root.ids.main_scroll.scroll_to(widget)

    def init_data_table(self):
        """Initializes live Excel grid table."""
        self.data_table = MDDataTable(
            use_pagination=True,
            rows_num=5,
            column_data=[
                ("#", dp(10)),
                ("Photo", dp(15)),
                ("Description", dp(30)),
                ("Price", dp(18)),
                ("MOQ", dp(12)),
                ("Qty/CTN", dp(18)),
                ("CBM/CTN", dp(18)),
                ("Shop", dp(22)),
                ("Phone", dp(22)),
                ("WeChat", dp(22)),
            ],
            row_data=[]
        )
        self.root.ids.table_container.clear_widgets()
        self.root.ids.table_container.add_widget(self.data_table)

    # -------------------------------------------------------------------------
    # NATIVE CAMERA INTENT (ITEM PHOTO & CARD OCR MODE)
    # -------------------------------------------------------------------------
    def launch_native_camera(self, mode="item"):
        """Launches native camera for item picture or business card OCR scan."""
        try:
            filename = f"IMG_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
            target_path = os.path.join(self.photos_dir, filename)

            if mode == "item":
                self.current_photo_path = target_path
                request_code = self.CAMERA_ITEM_REQUEST_CODE
            else:
                self.card_photo_path = target_path
                request_code = self.CAMERA_CARD_REQUEST_CODE

            if platform == 'android':
                from jnius import autoclass, cast
                
                StrictMode = autoclass('android.os.StrictMode')
                VmPolicyBuilder = autoclass('android.os.StrictMode$VmPolicy$Builder')
                StrictMode.setVmPolicy(VmPolicyBuilder().build())

                Intent = autoclass('android.content.Intent')
                MediaStore = autoclass('android.provider.MediaStore')
                File = autoclass('java.io.File')
                Uri = autoclass('android.net.Uri')
                PythonActivity = autoclass('org.kivy.android.PythonActivity')

                photo_file = File(target_path)
                photo_uri = Uri.fromFile(photo_file)
                parcelable_uri = cast('android.os.Parcelable', photo_uri)

                intent = Intent(MediaStore.ACTION_IMAGE_CAPTURE)
                intent.putExtra(MediaStore.EXTRA_OUTPUT, parcelable_uri)

                activity = PythonActivity.mActivity
                activity.startActivityForResult(intent, request_code)
            else:
                self.show_dialog("Desktop Mode", "Native camera intents are supported on Android devices.")
        except Exception as e:
            self.show_dialog("Camera Error", f"Could not launch native camera:\n{str(e)}")

    def on_activity_result(self, request_code, result_code, intent):
        """Processes activity results for item photo and business card OCR scan."""
        if request_code == self.CAMERA_ITEM_REQUEST_CODE:
            if result_code == -1 or os.path.exists(self.current_photo_path):
                self.scan_file_to_gallery(self.current_photo_path)
                self.root.ids.image_preview.source = self.current_photo_path
                self.root.ids.image_preview.reload()

        elif request_code == self.CAMERA_CARD_REQUEST_CODE:
            if result_code == -1 or os.path.exists(self.card_photo_path):
                self.scan_file_to_gallery(self.card_photo_path)
                self.process_card_ocr(self.card_photo_path)

    # -------------------------------------------------------------------------
    # BUSINESS CARD OCR SCANNER & PARSER (COMPATIBLE WITH ALL ANDROID VERSIONS)
    # -------------------------------------------------------------------------
    def process_card_ocr(self, image_path):
        """Extracts text from business card photo and auto-populates input fields."""
        extracted_text = ""

        # Attempt 1: Try Google Play Services ML Kit / Vision Recognizer
        if platform == 'android':
            try:
                from jnius import autoclass
                
                BitmapFactory = autoclass('android.graphics.BitmapFactory')
                TextRecognizerBuilder = autoclass('com.google.android.gms.vision.text.TextRecognizer$Builder')
                FrameBuilder = autoclass('com.google.android.gms.vision.Frame$Builder')
                PythonActivity = autoclass('org.kivy.android.PythonActivity')

                bitmap = BitmapFactory.decodeFile(image_path)
                detector = TextRecognizerBuilder(PythonActivity.mActivity).build()

                if detector.isOperational():
                    frame = FrameBuilder().setBitmap(bitmap).build()
                    items = detector.detect(frame)

                    lines = []
                    for i in range(items.size()):
                        item = items.valueAt(i)
                        lines.append(item.getValue())
                    extracted_text = "\n".join(lines)
            except Exception as e:
                print(f"Android Vision OCR API unavailable: {e}")

        # Attempt 2: Try pytesseract if installed via pip in Pydroid
        if not extracted_text:
            try:
                import pytesseract
                from PIL import Image as PILImage
                img = PILImage.open(image_path)
                extracted_text = pytesseract.image_to_string(img)
            except Exception as e:
                print(f"Pytesseract not installed/available: {e}")

        # Attempt 3: Native text extraction fallback if no engine installed
        if not extracted_text:
            self.show_dialog(
                "OCR Engine Missing", 
                "To enable automatic card scanning, please install Tesseract in Pydroid 3:\n\n1. Go to Pydroid Pip menu\n2. Install 'pytesseract'\n\nFor now, enter details manually."
            )
            return

        # Smart regex field extraction
        phone_match = re.search(r'(\+?\d{2,4}[\s-]?)?1[3-9]\d{9}|\b\d{3,4}[-.\s]?\d{7,8}\b', extracted_text)
        wechat_match = re.search(r'(?:WX|WeChat|微信|微信号)[:\s]*([a-zA-Z0-9_-]+)', extracted_text, re.IGNORECASE)
        qq_match = re.search(r'(?:QQ)[:\s]*([1-9][0-9]{4,10})', extracted_text, re.IGNORECASE)

        ids = self.root.ids
        parsed_fields = []

        if phone_match:
            ids.field_phone.text = phone_match.group(0).strip()
            parsed_fields.append("Phone")
        if wechat_match:
            ids.field_wechat.text = wechat_match.group(1).strip()
            parsed_fields.append("WeChat")
        if qq_match:
            ids.field_qq.text = qq_match.group(1).strip()
            parsed_fields.append("QQ")

        lines = [line.strip() for line in extracted_text.splitlines() if line.strip()]
        if lines:
            # Use top line as shop/booth description
            ids.field_shop.text = lines[0]
            parsed_fields.append("Shop Name")

        if parsed_fields:
            self.show_dialog("Card Parsed", f"Successfully extracted: {', '.join(parsed_fields)}")
        else:
            self.show_dialog("Scan Result", "Could not identify distinct phone or WeChat numbers. Check photo lighting and try again.")

    def scan_file_to_gallery(self, file_path):
        """Notifies Android Gallery indexer."""
        try:
            from jnius import autoclass
            MediaScannerConnection = autoclass('android.media.MediaScannerConnection')
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            MediaScannerConnection.scanFile(PythonActivity.mActivity, [file_path], None, None)
        except Exception:
            pass

    # -------------------------------------------------------------------------
    # FILE SELECTOR
    # -------------------------------------------------------------------------
    def open_file_browser_popup(self):
        """Opens pure Kivy file browser."""
        try:
            content = BoxLayout(orientation='vertical', spacing=10, padding=10)
            start_path = self.photos_dir if os.path.exists(self.photos_dir) else os.path.expanduser("~")
            
            filechooser_widget = FileChooserListView(
                path=start_path,
                filters=["*.png", "*.jpg", "*.jpeg", "*.PNG", "*.JPG", "*.JPEG"]
            )
            content.add_widget(filechooser_widget)

            btn_box = BoxLayout(orientation='horizontal', size_hint_y=None, height='48dp', spacing=10)
            select_btn = MDRaisedButton(
                text="Select",
                icon="check",
                on_release=lambda x: self.on_kivy_file_selected(filechooser_widget.selection)
            )
            cancel_btn = MDFlatButton(text="Cancel", on_release=lambda x: self.file_popup.dismiss())
            
            btn_box.add_widget(select_btn)
            btn_box.add_widget(cancel_btn)
            content.add_widget(btn_box)

            self.file_popup = Popup(
                title="Select Photo from Storage",
                content=content,
                size_hint=(0.95, 0.85),
                auto_dismiss=False
            )
            self.file_popup.open()
        except Exception as e:
            self.show_dialog("Picker Error", str(e))

    def on_kivy_file_selected(self, selection):
        if selection and len(selection) > 0:
            self.current_photo_path = selection[0]
            self.root.ids.image_preview.source = self.current_photo_path
            self.root.ids.image_preview.reload()
            if self.file_popup:
                self.file_popup.dismiss()

    # -------------------------------------------------------------------------
    # EXCEL LOGIC & CUSTOM FILE NAMING PROMPTS
    # -------------------------------------------------------------------------
    def save_quote(self):
        """Appends entry to list and updates active Excel sheet."""
        ids = self.root.ids
        desc = ids.field_desc.text.strip()
        
        if not desc:
            self.show_dialog("Error", "Please enter at least an Item Description.")
            return

        quote = {
            "photo": self.current_photo_path,
            "desc": desc,
            "price": ids.field_price.text,
            "moq": ids.field_moq.text,
            "qty_ctn": ids.field_qty_ctn.text,
            "cbm_ctn": ids.field_cbm_ctn.text,
            "shop": ids.field_shop.text,
            "phone": ids.field_phone.text,
            "wechat": ids.field_wechat.text,
            "qq": ids.field_qq.text,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M")
        }

        self.quotes.append(quote)
        
        row_num = str(len(self.quotes))
        has_img = "Yes" if quote["photo"] else "No"
        self.data_table.add_row((
            row_num,
            has_img,
            quote["desc"],
            quote["price"],
            quote["moq"],
            quote["qty_ctn"],
            quote["cbm_ctn"],
            quote["shop"],
            quote["phone"],
            quote["wechat"]
        ))

        if self.current_session_filename:
            self.export_to_excel(show_alert=False)
        self.clear_form()

    def clear_form(self):
        ids = self.root.ids
        for field in [ids.field_desc, ids.field_price, ids.field_moq, 
                      ids.field_qty_ctn, ids.field_cbm_ctn, ids.field_shop, 
                      ids.field_phone, ids.field_wechat, ids.field_qq]:
            field.text = ""
        ids.image_preview.source = ""
        self.current_photo_path = ""

    def prompt_excel_filename(self):
        """Prompts user to name/rename Excel file when clicking Save."""
        if not self.quotes:
            self.show_dialog("Export", "No items added to sheet yet.")
            return

        default_name = self.current_session_filename.replace(".xlsx", "") if self.current_session_filename else f"Quotes_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        input_field = MDTextField(text=default_name, hint_text="Enter Excel File Name")

        self.naming_dialog = MDDialog(
            title="Name Excel File",
            type="custom",
            content_cls=input_field,
            buttons=[
                MDFlatButton(text="Cancel", on_release=lambda x: self.naming_dialog.dismiss()),
                MDRaisedButton(
                    text="Save Sheet",
                    on_release=lambda x: self.confirm_file_save(input_field.text.strip())
                ),
            ],
        )
        self.naming_dialog.open()

    def confirm_file_save(self, custom_name):
        """Handles confirmation from file naming dialog."""
        if not custom_name:
            custom_name = f"Quotes_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
        if not custom_name.endswith(".xlsx"):
            custom_name += ".xlsx"

        self.current_session_filename = custom_name
        if self.naming_dialog:
            self.naming_dialog.dismiss()
            
        self.export_to_excel(show_alert=True)

    def export_to_excel(self, show_alert=True):
        """Exports data into designated Excel sheet inside same folder as photos."""
        if not self.quotes:
            if show_alert:
                self.show_dialog("Export", "No items added yet.")
            return

        if not self.current_session_filename:
            self.prompt_excel_filename()
            return

        try:
            from PIL import Image as PILImage

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Quote Sheet"

            headers = ["Photo", "Description", "Price ($)", "MOQ", "Qty/CTN", "CBM/CTN", "Shop No.", "Phone", "WeChat", "QQ", "Date Added"]
            ws.append(headers)

            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            thin_border = Border(left=Side(style='thin', color='D9D9D9'),
                                 right=Side(style='thin', color='D9D9D9'),
                                 top=Side(style='thin', color='D9D9D9'),
                                 bottom=Side(style='thin', color='D9D9D9'))

            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[1].height = 28
            ws.column_dimensions['A'].width = 18

            for row_idx, q in enumerate(self.quotes, start=2):
                ws.row_dimensions[row_idx].height = 85
                
                row_data = [
                    "",
                    q["desc"],
                    float(q["price"]) if q["price"] else "",
                    int(q["moq"]) if q["moq"] else "",
                    int(q["qty_ctn"]) if q["qty_ctn"] else "",
                    float(q["cbm_ctn"]) if q["cbm_ctn"] else "",
                    q["shop"],
                    q["phone"],
                    q["wechat"],
                    q["qq"],
                    q["timestamp"]
                ]
                
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                photo_path = q["photo"]
                if photo_path and os.path.exists(photo_path):
                    try:
                        pil_img = PILImage.open(photo_path)
                        
                        try:
                            from PIL import ImageOps
                            pil_img = ImageOps.exif_transpose(pil_img)
                        except Exception:
                            pass

                        if pil_img.mode in ("RGBA", "P"):
                            pil_img = pil_img.convert("RGB")

                        # Crisp ~800px max dimension embedding
                        max_dimension = 800
                        width, height = pil_img.size
                        if width > max_dimension or height > max_dimension:
                            if width > height:
                                new_w = max_dimension
                                new_h = int((max_dimension / float(width)) * height)
                            else:
                                new_h = max_dimension
                                new_w = int((max_dimension / float(height)) * width)
                            pil_img = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)

                        thumb_name = f"excel_{os.path.basename(photo_path)}.jpg"
                        temp_resized_path = os.path.join(self.photos_dir, thumb_name)
                        
                        pil_img.save(temp_resized_path, "JPEG", quality=85, optimize=True)

                        img = OpenpyxlImage(temp_resized_path)
                        img.width = 110
                        img.height = 100
                        ws.add_image(img, f"A{row_idx}")
                    except Exception as img_err:
                        print(f"Image embed error: {img_err}")

            for col in range(2, len(headers) + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 16

            out_path = os.path.join(self.excel_dir, self.current_session_filename)
            wb.save(out_path)

            if show_alert:
                self.show_dialog("Saved Successfully", f"Excel Sheet saved to:\n{out_path}")

        except Exception as e:
            self.show_dialog("Export Error", str(e))

    def finish_session_prompt(self):
        """Prompts user for filename before completing session."""
        if not self.quotes:
            self.show_dialog("Session", "Current session is already empty.")
            return

        if not self.current_session_filename:
            self.prompt_excel_filename()
        else:
            self.export_to_excel(show_alert=False)
            old_file = self.current_session_filename

            self.quotes = []
            self.current_session_filename = ""
            self.init_data_table()
            
            self.show_dialog("Session Complete", f"Saved file to:\n{os.path.join(self.excel_dir, old_file)}\n\nReady for a new quote session!")

    def show_dialog(self, title, text):
        """Displays dialog alert."""
        dialog = MDDialog(
            title=title,
            text=text,
            size_hint=(0.85, None),
            buttons=[MDFlatButton(text="OK", on_release=lambda x: dialog.dismiss())]
        )
        dialog.open()

if __name__ == "__main__":
    QuoteApp().run()